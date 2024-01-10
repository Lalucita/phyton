import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import os
from openpyxl.styles import NamedStyle, numbers
import openpyxl



from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def cambiar_formato_xlsx(xlsx_path):
    # Cargar el archivo Excel existente
    book = load_workbook(xlsx_path)
    hoja = book['Hoja1']
    

    # Definir colores y estilos solo para los encabezados
    borde_negro = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))     

    # Aplicar bordes a todas las celdas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column):
        for celda in fila:
            # Aplicar borde a las celdas
            celda.border = borde_negro

    # Aplicar estilos a las celdas de la primera fila (encabezados)
    for celda in hoja[1]:
        # Crear una copia del estilo de fuente y relleno
        nuevo_estilo_fuente = Font(bold=True, color='FFFFFF')
        nuevo_estilo_relleno = PatternFill(start_color='26B461', end_color='26B461', fill_type='solid')

        # Asignar el nuevo estilo a la celda
        celda.font = nuevo_estilo_fuente
        celda.fill = nuevo_estilo_relleno

    # Guardar los cambios en el archivo Excel
    book.save(xlsx_path)
    print(f'Formato del archivo Excel modificado: {xlsx_path}')


def convertir_csv_a_xlsx(csv_path, xlsx_path):
    # Leer el archivo CSV (separado por comas por defecto)
    df = pd.read_csv(csv_path)

    # Crear un nuevo archivo Excel (xlsx) con openpyxl
    writer = pd.ExcelWriter(xlsx_path, engine='openpyxl')
    writer.book = Workbook()
    df.to_excel(writer, sheet_name='Sheet', index=False)
    hoja = writer.sheets['Sheet']

    # Definir colores y estilos solo para los encabezados
    borde_negro = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))     


    # Aplicar bordes a todas las celdas de la hoja
    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column):
        for celda in fila:
            # Aplicar borde a las celdas
            celda.border = borde_negro

    # Quitar decimales de las columnas numéricas
    for columna in df.columns:
        if pd.api.types.is_numeric_dtype(df[columna]) and columna != 'NroCliente':
            df[columna] = df[columna].apply(lambda x: round(x) if x == x else x)  # round si no es NaN

    # Guardar el DataFrame actualizado en el archivo Excel
    df.to_excel(writer, sheet_name='Sheet', index=False)
    
    
    # Aplicar estilos a las celdas de la primera fila (encabezados)
    for celda in hoja[1]:
        # Crear una copia del estilo de fuente y relleno
        nuevo_estilo_fuente = Font(bold=True, color='FFFFFF')
        nuevo_estilo_relleno = PatternFill(start_color='26B461', end_color='26B461', fill_type='solid')

        # Asignar el nuevo estilo a la celda
        celda.font = nuevo_estilo_fuente
        celda.fill = nuevo_estilo_relleno

    # Cerrar el writer
    writer.save()
    print(f'Archivo Excel creado: {xlsx_path}')


def convertir_y_eliminar_csv_en_directorio(directorio_csv, directorio_excel):
    # Asegurar que el directorio de salida exista
    os.makedirs(directorio_excel, exist_ok=True)

    # Listar todos los archivos CSV en el directorio
    archivos_csv = [archivo for archivo in os.listdir(directorio_csv) if archivo.endswith('.csv')]

    # Iterar sobre cada archivo CSV y convertirlo
    for archivo_csv in archivos_csv:
        ruta_csv = os.path.join(directorio_csv, archivo_csv)
        ruta_excel = os.path.join(directorio_excel, archivo_csv.replace('.csv', '.xlsx'))

        # Llamar a la función que creamos anteriormente
        convertir_csv_a_xlsx(ruta_csv, ruta_excel)

        # Eliminar el archivo original CSV
        os.remove(ruta_csv)

# Ejemplo de uso
directorio_csv = r'D:\Lucia\oportunidades_de_negocio'
directorio_excel = r'D:\Lucia\oportunidades_de_negocio\excel_con_formato'

#convertir_y_eliminar_csv_en_directorio(directorio_csv, directorio_excel)

cambiar_formato_xlsx("D:/Lucia/2024_solicitudes/samuel_ene/microcreditos.xlsx")
